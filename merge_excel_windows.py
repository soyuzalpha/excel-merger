from openpyxl import load_workbook, Workbook
import os
from datetime import datetime
import copy
import sys

def copy_cell(source_cell, target_cell):
    target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def get_current_folder():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

def merge_folder():
    folder_path = get_current_folder()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(folder_path, f"merged_{timestamp}.xlsx")

    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") and not filename.startswith("~$") and not 
filename.startswith("merged_"):

            file_path = os.path.join(folder_path, filename)

            try:
                wb = load_workbook(file_path, data_only=False)

                for sheet_name in wb.sheetnames:
                    source_sheet = wb[sheet_name]

                    new_sheet_name = f"{os.path.splitext(filename)[0]}_{sheet_name}"
                    new_sheet_name = new_sheet_name[:31]

                    counter = 1
                    base_name = new_sheet_name
                    while new_sheet_name in new_wb.sheetnames:
                        new_sheet_name = f"{base_name[:28]}_{counter}"
                        counter += 1

                    target_sheet = new_wb.create_sheet(title=new_sheet_name)

                    for row in source_sheet.iter_rows():
                        for cell in row:
                            target_cell = target_sheet[cell.coordinate]
                            copy_cell(cell, target_cell)

                    # Copy column width
                    for col in source_sheet.column_dimensions:
                        target_sheet.column_dimensions[col].width = \
                            source_sheet.column_dimensions[col].width

                    # Copy merged cells
                    for merged_cell in source_sheet.merged_cells.ranges:
                        target_sheet.merge_cells(str(merged_cell))

            except Exception as e:
                print(f"Gagal baca {filename}: {e}")

    new_wb.save(output_file)
    print(f"Merge selesai! File dibuat: {output_file}")

if __name__ == "__main__":
    merge_folder()

